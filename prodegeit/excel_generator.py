"""
ProDegeit Project - Excel Generator
Creates comprehensive Excel reports
"""

from typing import Dict, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from data_models import ACTIVITIES, RESOURCES, ALL_SKILLS


class ExcelGenerator:
    """Generates Excel workbooks for project data"""
    
    def __init__(self):
        """Initialize Excel generator"""
        pass
    
    def generate_resource_workbook(self, output_path: str):
        """
        Generate resource data workbook
        
        Args:
            output_path: Path to save Excel file
        """
        print(f"\nGenerating Resource Excel: {output_path}")
        
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Sheet 1: Resource Master Data
        self._create_resource_sheet(wb)
        
        # Sheet 2: Skill Matrix
        self._create_skill_matrix_sheet(wb)
        
        # Sheet 3: Availability Calendar
        self._create_availability_sheet(wb)
        
        wb.save(output_path)
        print(f"✓ Resource workbook generated")
    
    def generate_allocation_workbook(self, allocation_results: Dict, output_path: str):
        """
        Generate allocation results workbook
        
        Args:
            allocation_results: Results from resource allocator
            output_path: Path to save Excel file
        """
        print(f"\nGenerating Allocation Excel: {output_path}")
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Sheet 1: Activity List
        self._create_activity_sheet(wb, allocation_results)
        
        # Sheet 2: Allocation Matrix
        self._create_allocation_matrix_sheet(wb, allocation_results)
        
        # Sheet 3: Resource Utilization
        self._create_utilization_sheet(wb, allocation_results)
        
        # Sheet 4: Cost Summary
        self._create_cost_sheet(wb, allocation_results)
        
        wb.save(output_path)
        print(f"✓ Allocation workbook generated")
    
    def _create_resource_sheet(self, wb):
        """Create resource master data sheet"""
        ws = wb.create_sheet("Resources")
        
        # Headers
        headers = ["Name", "Cost/Hour (€)", "Availability %", "Start Week", 
                  "Vacation Weeks", "Core Team"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data
        for row, resource in enumerate(RESOURCES, start=2):
            ws.cell(row, 1, resource.name)
            ws.cell(row, 2, resource.cost_per_hour)
            ws.cell(row, 3, f"{resource.availability_pct*100}%")
            ws.cell(row, 4, resource.start_week)
            ws.cell(row, 5, ", ".join(map(str, resource.vacation_weeks)) if resource.vacation_weeks else "-")
            ws.cell(row, 6, "Yes" if resource.is_core_team else "No")
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_skill_matrix_sheet(self, wb):
        """Create skill matrix heat map"""
        ws = wb.create_sheet("Skill Matrix")
        
        # Headers
        ws.cell(1, 1, "Resource")
        for col, skill in enumerate(ALL_SKILLS, start=2):
            cell = ws.cell(1, col, skill)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data with conditional formatting
        for row, resource in enumerate(RESOURCES, start=2):
            ws.cell(row, 1, resource.name).font = Font(bold=True)
            
            for col, skill in enumerate(ALL_SKILLS, start=2):
                level = resource.skills.get(skill, 0)
                cell = ws.cell(row, col, level if level > 0 else "-")
                
                # Color code by skill level
                if level == 0:
                    color = "FFFFFF"
                elif level <= 2:
                    color = "FFF2CC"
                elif level <= 4:
                    color = "FFD966"
                else:
                    color = "92D050"
                
                cell.fill = PatternFill(start_color=color, fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        # Auto-size
        ws.column_dimensions['A'].width = 15
        for col in range(2, len(ALL_SKILLS) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_availability_sheet(self, wb):
        """Create availability calendar"""
        ws = wb.create_sheet("Availability")
        
        # Headers
        ws.cell(1, 1, "Resource")
        for week in range(1, 13):  # 12 week project
            cell = ws.cell(1, week + 1, f"Week {week}")
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row, resource in enumerate(RESOURCES, start=2):
            ws.cell(row, 1, resource.name)
            
            for week in range(1, 13):
                cell = ws.cell(row, week + 1)
                
                if resource.is_available(week):
                    cell.value = f"{int(resource.availability_pct*100)}%"
                    cell.fill = PatternFill(start_color="C6E0B4", fill_type="solid")
                else:
                    cell.value = "N/A"
                    cell.fill = PatternFill(start_color="F4B084", fill_type="solid")
                
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(size=9)
        
        ws.column_dimensions['A'].width = 15
        for col in range(2, 14):
            ws.column_dimensions[get_column_letter(col)].width = 8
    
    def _create_activity_sheet(self, wb, results):
        """Create activity list sheet"""
        ws = wb.create_sheet("Activities")
        
        # Headers
        headers = ["ID", "Activity Name", "Duration (days)", "Predecessors", 
                  "Start Date", "End Date", "Allocated Resources"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data
        for activity in ACTIVITIES:
            row = activity.id + 1
            schedule = results['schedule'].get(activity.id, {})
            allocated = results['allocation_map'].get(activity.id, [])
            
            ws.cell(row, 1, activity.id)
            ws.cell(row, 2, activity.name)
            ws.cell(row, 3, activity.duration_days)
            ws.cell(row, 4, ", ".join(map(str, activity.predecessors)) if activity.predecessors else "-")
            ws.cell(row, 5, schedule.get('start', '-').strftime("%Y-%m-%d") if schedule.get('start') else "-")
            ws.cell(row, 6, schedule.get('end', '-').strftime("%Y-%m-%d") if schedule.get('end') else "-")
            ws.cell(row, 7, ", ".join(allocated) if allocated else "-")
        
        # Auto-size
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20 if col == 2 else 15
    
    def _create_allocation_matrix_sheet(self, wb, results):
        """Create resource allocation matrix"""
        ws = wb.create_sheet("Allocation Matrix")
        
        # Headers - Activities across top
        ws.cell(1, 1, "Resource \\ Activity")
        for col, activity in enumerate(ACTIVITIES, start=2):
            cell = ws.cell(1, col, f"A{activity.id}")
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal="center", text_rotation=90)
            ws.column_dimensions[get_column_letter(col)].width = 4
        
        # Add totals column
        ws.cell(1, len(ACTIVITIES) + 2, "Total Tasks")
        ws.cell(1, len(ACTIVITIES) + 3, "Total Hours")
        
        # Resource rows
        for row, resource in enumerate(RESOURCES, start=2):
            ws.cell(row, 1, resource.name).font = Font(bold=True)
            
            task_count = 0
            for col, activity in enumerate(ACTIVITIES, start=2):
                allocated = results['allocation_map'].get(activity.id, [])
                
                if resource.name in allocated:
                    cell = ws.cell(row, col, "✓")
                    cell.fill = PatternFill(start_color="92D050", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                    task_count += 1
                else:
                    ws.cell(row, col, "")
            
            # Totals
            util = results['resource_utilization'].get(resource.name, {})
            ws.cell(row, len(ACTIVITIES) + 2, task_count)
            ws.cell(row, len(ACTIVITIES) + 3, f"{util.get('hours', 0):.0f}")
        
        ws.column_dimensions['A'].width = 15
    
    def _create_utilization_sheet(self, wb, results):
        """Create resource utilization summary"""
        ws = wb.create_sheet("Utilization")
        
        # Headers
        headers = ["Resource", "Total Hours", "Total Cost (€)", "Number of Tasks"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data
        row = 2
        for resource_name, util in sorted(
            results['resource_utilization'].items(),
            key=lambda x: x[1]['cost'],
            reverse=True
        ):
            ws.cell(row, 1, resource_name)
            ws.cell(row, 2, f"{util['hours']:.1f}")
            ws.cell(row, 3, f"€{util['cost']:,.2f}")
            ws.cell(row, 4, util['tasks'])
            row += 1
        
        # Total row
        row += 1
        ws.cell(row, 1, "TOTAL").font = Font(bold=True)
        ws.cell(row, 3, f"€{results['estimated_cost']:,.2f}").font = Font(bold=True)
        
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_cost_sheet(self, wb, results):
        """Create cost breakdown sheet"""
        ws = wb.create_sheet("Cost Analysis")
        
        # Headers
        ws.cell(1, 1, "Activity ID").font = Font(bold=True)
        ws.cell(1, 2, "Activity Name").font = Font(bold=True)
        ws.cell(1, 3, "Cost (€)").font = Font(bold=True)
        
        # Data
        row = 2
        for activity in ACTIVITIES:
            cost = results['activity_costs'].get(activity.id, 0)
            ws.cell(row, 1, activity.id)
            ws.cell(row, 2, activity.name)
            ws.cell(row, 3, f"€{cost:,.2f}")
            row += 1
        
        # Core team cost
        row += 1
        ws.cell(row, 2, "Core Team (Fixed)").font = Font(bold=True)
        ws.cell(row, 3, f"€{results['core_team_cost']:,.2f}").font = Font(bold=True)
        
        # Total
        row += 1
        ws.cell(row, 2, "TOTAL PROJECT COST").font = Font(bold=True, size=12)
        ws.cell(row, 3, f"€{results['estimated_cost']:,.2f}").font = Font(bold=True, size=12)
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 18


if __name__ == "__main__":
    print("Testing Excel Generator...")
    
    generator = ExcelGenerator()
    
    # Generate resource workbook
    generator.generate_resource_workbook("output/ProDegeit_Resources.xlsx")
    
    print("\n✓ Excel generator test complete")
